from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from .models import Produto, Categoria, Fornecedor, MovimentacaoEstoque, Venda, ItemVenda
from .forms import ProdutoForm, MovimentacaoForm, VendaForm, ItemVendaFormSet
import json


@login_required
def dashboard(request):
    """Dashboard principal do sistema"""
    # Estatísticas gerais
    total_produtos = Produto.objects.filter(ativo=True).count()
    total_categorias = Categoria.objects.filter(ativa=True).count()
    total_fornecedores = Fornecedor.objects.filter(ativo=True).count()
    
    # Produtos com estoque baixo
    produtos_estoque_baixo = Produto.objects.filter(
        ativo=True,
        estoque_atual__lte=F('estoque_minimo')
    ).count()
    
    # Valor total do estoque
    produtos_ativos = Produto.objects.filter(ativo=True)
    valor_total_estoque = sum(produto.valor_total_estoque for produto in produtos_ativos)
    
    # Movimentações recentes
    movimentacoes_recentes = MovimentacaoEstoque.objects.select_related(
        'produto', 'usuario'
    ).order_by('-data_movimentacao')[:10]
    
    # Vendas do mês
    mes_atual = timezone.now().month
    ano_atual = timezone.now().year
    vendas_mes = Venda.objects.filter(
        data_venda__month=mes_atual,
        data_venda__year=ano_atual,
        status='CONCLUIDA'
    ).aggregate(
        total_vendas=Count('id'),
        total_valor=Sum('total')
    )
    
    # Produtos mais vendidos (últimos 30 dias)
    data_limite = timezone.now() - timezone.timedelta(days=30)
    produtos_mais_vendidos = ItemVenda.objects.filter(
        venda__data_venda__gte=data_limite,
        venda__status='CONCLUIDA'
    ).values('produto__nome').annotate(
        total_vendido=Sum('quantidade')
    ).order_by('-total_vendido')[:5]
    
    context = {
        'total_produtos': total_produtos,
        'total_categorias': total_categorias,
        'total_fornecedores': total_fornecedores,
        'produtos_estoque_baixo': produtos_estoque_baixo,
        'valor_total_estoque': valor_total_estoque,
        'movimentacoes_recentes': movimentacoes_recentes,
        'vendas_mes': vendas_mes,
        'produtos_mais_vendidos': produtos_mais_vendidos,
    }
    
    return render(request, 'estoque/dashboard.html', context)


@login_required
def lista_produtos(request):
    """Lista todos os produtos com filtros e busca"""
    produtos = Produto.objects.select_related('categoria', 'fornecedor').filter(ativo=True)
    
    # Filtros
    categoria_id = request.GET.get('categoria')
    fornecedor_id = request.GET.get('fornecedor')
    status_estoque = request.GET.get('status_estoque')
    busca = request.GET.get('busca')
    
    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)
    
    if fornecedor_id:
        produtos = produtos.filter(fornecedor_id=fornecedor_id)
    
    if status_estoque:
        if status_estoque == 'baixo':
            produtos = produtos.filter(estoque_atual__lte=F('estoque_minimo'))
        elif status_estoque == 'sem_estoque':
            produtos = produtos.filter(estoque_atual=0)
    
    if busca:
        produtos = produtos.filter(
            Q(nome__icontains=busca) | 
            Q(codigo__icontains=busca) |
            Q(descricao__icontains=busca)
        )
    
    # Paginação
    paginator = Paginator(produtos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Dados para filtros
    categorias = Categoria.objects.filter(ativa=True)
    fornecedores = Fornecedor.objects.filter(ativo=True)
    
    context = {
        'page_obj': page_obj,
        'categorias': categorias,
        'fornecedores': fornecedores,
        'filtros': {
            'categoria': categoria_id,
            'fornecedor': fornecedor_id,
            'status_estoque': status_estoque,
            'busca': busca,
        }
    }
    
    return render(request, 'estoque/lista_produtos.html', context)


@login_required
def detalhes_produto(request, produto_id):
    """Detalhes de um produto específico"""
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Histórico de movimentações
    movimentacoes = MovimentacaoEstoque.objects.filter(
        produto=produto
    ).select_related('usuario').order_by('-data_movimentacao')[:20]
    
    context = {
        'produto': produto,
        'movimentacoes': movimentacoes,
    }
    
    return render(request, 'estoque/detalhes_produto.html', context)


@login_required
def adicionar_produto(request):
    """Adicionar novo produto"""
    if request.method == 'POST':
        form = ProdutoForm(request.POST)
        if form.is_valid():
            produto = form.save(commit=False)
            produto.save()
            messages.success(request, 'Produto adicionado com sucesso!')
            return redirect('estoque:detalhes_produto', produto_id=produto.id)
    else:
        form = ProdutoForm()
    
    return render(request, 'estoque/adicionar_produto.html', {'form': form})


@login_required
def editar_produto(request, produto_id):
    """Editar produto existente"""
    produto = get_object_or_404(Produto, id=produto_id)
    
    if request.method == 'POST':
        form = ProdutoForm(request.POST, instance=produto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produto atualizado com sucesso!')
            return redirect('estoque:detalhes_produto', produto_id=produto.id)
    else:
        form = ProdutoForm(instance=produto)
    
    return render(request, 'estoque/editar_produto.html', {'form': form, 'produto': produto})


@login_required
def movimentacao_estoque(request):
    """Realizar movimentação de estoque"""
    if request.method == 'POST':
        form = MovimentacaoForm(request.POST)
        if form.is_valid():
            movimentacao = form.save(commit=False)
            movimentacao.usuario = request.user
            movimentacao.save()
            
            messages.success(request, 'Movimentação realizada com sucesso!')
            return redirect('estoque:dashboard')
    else:
        form = MovimentacaoForm()
    
    return render(request, 'estoque/movimentacao_estoque.html', {'form': form})


@login_required
def nova_venda(request):
    """Criar nova venda"""
    if request.method == 'POST':
        form = VendaForm(request.POST)
        formset = ItemVendaFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            venda = form.save(commit=False)
            venda.vendedor = request.user
            venda.numero_venda = f"VENDA-{timezone.now().strftime('%Y%m%d%H%M%S')}"
            venda.save()
            
            # Salvar itens da venda
            for item_form in formset:
                if item_form.cleaned_data:
                    item = item_form.save(commit=False)
                    item.venda = venda
                    item.save()
                    
                    # Criar movimentação de saída
                    MovimentacaoEstoque.objects.create(
                        produto=item.produto,
                        tipo='SAIDA',
                        quantidade=item.quantidade,
                        usuario=request.user,
                        observacoes=f"Venda #{venda.numero_venda}"
                    )
            
            # Calcular total da venda
            venda.total = sum(item.subtotal for item in venda.itens.all())
            venda.save()
            
            messages.success(request, f'Venda #{venda.numero_venda} criada com sucesso!')
            return redirect('estoque:detalhes_venda', venda_id=venda.id)
    else:
        form = VendaForm()
        formset = ItemVendaFormSet()
    
    return render(request, 'estoque/nova_venda.html', {'form': form, 'formset': formset})


@login_required
def detalhes_venda(request, venda_id):
    """Detalhes de uma venda"""
    venda = get_object_or_404(Venda, id=venda_id)
    itens = venda.itens.select_related('produto').all()
    
    context = {
        'venda': venda,
        'itens': itens,
    }
    
    return render(request, 'estoque/detalhes_venda.html', context)


@login_required
def buscar_produto_ajax(request):
    """Busca produtos via AJAX para autocomplete"""
    termo = request.GET.get('term', '')
    
    if len(termo) < 2:
        return JsonResponse({'produtos': []})
    
    produtos = Produto.objects.filter(
        Q(nome__icontains=termo) | Q(codigo__icontains=termo),
        ativo=True
    )[:10]
    
    resultados = []
    for produto in produtos:
        resultados.append({
            'id': produto.id,
            'codigo': produto.codigo,
            'nome': produto.nome,
            'preco_venda': float(produto.preco_venda),
            'estoque_atual': produto.estoque_atual,
            'unidade': produto.get_unidade_display(),
        })
    
    return JsonResponse({'produtos': resultados})


@login_required
def relatorios(request):
    """Página de relatórios com gráficos e estatísticas detalhadas"""
    from django.db.models import Sum, Count, Avg
    from datetime import datetime, timedelta
    
    # Dados básicos de estoque
    produtos_estoque_baixo = Produto.objects.filter(
        ativo=True,
        estoque_atual__lte=F('estoque_minimo')
    ).count()
    
    produtos_sem_estoque = Produto.objects.filter(
        ativo=True,
        estoque_atual=0
    ).count()
    
    total_produtos = Produto.objects.filter(ativo=True).count()
    
    # Valor total do estoque
    valor_total_estoque = sum(
        produto.estoque_atual * produto.preco_custo 
        for produto in Produto.objects.filter(ativo=True)
    )
    
    # Movimentações dos últimos 30 dias
    data_30_dias_atras = timezone.now() - timedelta(days=30)
    movimentacoes_30_dias = MovimentacaoEstoque.objects.filter(
        data_movimentacao__gte=data_30_dias_atras
    ).count()
    
    # Vendas por mês (últimos 6 meses)
    vendas_por_mes = []
    for i in range(6):
        data = timezone.now() - timedelta(days=30*i)
        mes = data.month
        ano = data.year
        
        total_vendas = Venda.objects.filter(
            data_venda__month=mes,
            data_venda__year=ano,
            status='CONCLUIDA'
        ).count()
        
        vendas_por_mes.append({
            'mes': data.strftime('%m/%Y'),
            'total': total_vendas
        })
    
    vendas_por_mes.reverse()
    
    # Movimentações por tipo (últimos 30 dias)
    movimentacoes_entrada = MovimentacaoEstoque.objects.filter(
        tipo='ENTRADA',
        data_movimentacao__gte=data_30_dias_atras
    ).count()
    
    movimentacoes_saida = MovimentacaoEstoque.objects.filter(
        tipo='SAIDA',
        data_movimentacao__gte=data_30_dias_atras
    ).count()
    
    # Produtos por categoria
    produtos_por_categoria = []
    for categoria in Categoria.objects.filter(ativa=True):
        count = Produto.objects.filter(categoria=categoria, ativo=True).count()
        if count > 0:
            produtos_por_categoria.append({
                'categoria': categoria.nome,
                'quantidade': count
            })
    
    # Top 10 produtos com mais movimentações (últimos 30 dias)
    top_produtos_movimentacao = MovimentacaoEstoque.objects.filter(
        data_movimentacao__gte=data_30_dias_atras
    ).values('produto__nome').annotate(
        total_movimentacoes=Count('id')
    ).order_by('-total_movimentacoes')[:10]
    
    # Produtos com estoque baixo (detalhado)
    produtos_estoque_baixo_detalhado = Produto.objects.filter(
        ativo=True,
        estoque_atual__lte=F('estoque_minimo')
    ).select_related('categoria', 'fornecedor')[:10]
    
    # Movimentações por dia (últimos 7 dias)
    movimentacoes_por_dia = []
    for i in range(7):
        data = timezone.now() - timedelta(days=i)
        entrada = MovimentacaoEstoque.objects.filter(
            tipo='ENTRADA',
            data_movimentacao__date=data.date()
        ).count()
        saida = MovimentacaoEstoque.objects.filter(
            tipo='SAIDA',
            data_movimentacao__date=data.date()
        ).count()
        
        movimentacoes_por_dia.append({
            'dia': data.strftime('%d/%m'),
            'entrada': entrada,
            'saida': saida
        })
    
    movimentacoes_por_dia.reverse()
    
    # Fornecedores com mais produtos
    fornecedores_produtos = Fornecedor.objects.filter(ativo=True).annotate(
        total_produtos=Count('produto')
    ).order_by('-total_produtos')[:5]
    
    context = {
        'produtos_estoque_baixo': produtos_estoque_baixo,
        'produtos_sem_estoque': produtos_sem_estoque,
        'total_produtos': total_produtos,
        'valor_total_estoque': valor_total_estoque,
        'movimentacoes_30_dias': movimentacoes_30_dias,
        'vendas_por_mes': vendas_por_mes,
        'movimentacoes_entrada': movimentacoes_entrada,
        'movimentacoes_saida': movimentacoes_saida,
        'produtos_por_categoria': produtos_por_categoria,
        'top_produtos_movimentacao': top_produtos_movimentacao,
        'produtos_estoque_baixo_detalhado': produtos_estoque_baixo_detalhado,
        'movimentacoes_por_dia': movimentacoes_por_dia,
        'fornecedores_produtos': fornecedores_produtos,
    }
    
    return render(request, 'estoque/relatorios.html', context)


@login_required
def relatorio_produtos_pdf(request):
    """Gerar relatório de produtos em PDF"""
    from django.http import HttpResponse
    from django.template.loader import get_template
    from django.template import Context
    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Criar buffer para o PDF
    buffer = io.BytesIO()
    
    # Criar documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=TA_CENTER)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=12, spaceAfter=12)
    
    # Conteúdo do PDF
    story = []
    
    # Título
    story.append(Paragraph("Relatório de Produtos - Sistema de Estoque", title_style))
    story.append(Spacer(1, 20))
    
    # Dados dos produtos
    produtos = Produto.objects.filter(ativo=True).select_related('categoria', 'fornecedor')
    
    # Tabela de produtos
    data = [['Código', 'Nome', 'Categoria', 'Estoque', 'Preço', 'Status']]
    
    for produto in produtos:
        status = "Sem Estoque" if produto.estoque_atual == 0 else "Baixo" if produto.estoque_atual <= produto.estoque_minimo else "OK"
        data.append([
            produto.codigo,
            produto.nome[:30],  # Limitar tamanho
            produto.categoria.nome,
            f"{produto.estoque_atual}",
            f"R$ {produto.preco_custo:.2f}",
            status
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    
    # Obter conteúdo do buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Criar resposta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_produtos.pdf"'
    response.write(pdf_content)
    
    return response


@login_required
def relatorio_movimentacoes_excel(request):
    """Gerar relatório de movimentações em Excel"""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timedelta
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações de Estoque"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = ['Data', 'Produto', 'Tipo', 'Quantidade', 'Usuário', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Dados das movimentações (últimos 30 dias)
    data_30_dias_atras = timezone.now() - timedelta(days=30)
    movimentacoes = MovimentacaoEstoque.objects.filter(
        data_movimentacao__gte=data_30_dias_atras
    ).select_related('produto', 'usuario').order_by('-data_movimentacao')
    
    # Adicionar dados
    for row, movimentacao in enumerate(movimentacoes, 2):
        ws.cell(row=row, column=1, value=movimentacao.data_movimentacao.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=movimentacao.produto.nome)
        ws.cell(row=row, column=3, value=movimentacao.tipo)
        ws.cell(row=row, column=4, value=movimentacao.quantidade)
        ws.cell(row=row, column=5, value=movimentacao.usuario.username)
        ws.cell(row=row, column=6, value=movimentacao.observacoes)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Criar resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_movimentacoes.xlsx"'
    
    # Salvar workbook
    wb.save(response)
    
    return response


@login_required
def exportar_relatorio_excel(request):
    """Exportar relatórios em Excel com filtro de datas"""
    from django.http import HttpResponse
    from django.contrib import messages
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from .forms import RelatorioDataForm
    
    if request.method == 'POST':
        form = RelatorioDataForm(request.POST)
        if form.is_valid():
            data_inicio = form.cleaned_data['data_inicio']
            data_fim = form.cleaned_data['data_fim']
            tipo_relatorio = form.cleaned_data['tipo_relatorio']
            
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            title_font = Font(bold=True, size=14, color="2E86AB")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if tipo_relatorio == 'movimentacoes':
                ws.title = "Movimentações de Estoque"
                
                # Título
                ws.merge_cells('A1:F1')
                ws['A1'] = f"Relatório de Movimentações de Estoque - {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
                ws['A1'].font = title_font
                ws['A1'].alignment = center_alignment
                
                # Cabeçalhos
                headers = ['Data/Hora', 'Produto', 'Categoria', 'Tipo', 'Quantidade', 'Usuário', 'Observações']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                
                # Dados das movimentações
                movimentacoes = MovimentacaoEstoque.objects.filter(
                    data_movimentacao__date__gte=data_inicio,
                    data_movimentacao__date__lte=data_fim
                ).select_related('produto', 'produto__categoria', 'usuario').order_by('-data_movimentacao')
                
                for row, movimentacao in enumerate(movimentacoes, 4):
                    ws.cell(row=row, column=1, value=movimentacao.data_movimentacao.strftime('%d/%m/%Y %H:%M'))
                    ws.cell(row=row, column=2, value=movimentacao.produto.nome)
                    ws.cell(row=row, column=3, value=movimentacao.produto.categoria.nome)
                    ws.cell(row=row, column=4, value=movimentacao.tipo)
                    ws.cell(row=row, column=5, value=movimentacao.quantidade)
                    ws.cell(row=row, column=6, value=movimentacao.usuario.username)
                    ws.cell(row=row, column=7, value=movimentacao.observacoes)
                    
                    # Aplicar bordas
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).border = border
                
                # Ajustar largura das colunas
                column_widths = [20, 30, 20, 15, 12, 20, 40]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
                
                filename = f"movimentacoes_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
                
            elif tipo_relatorio == 'produtos':
                ws.title = "Produtos"
                
                # Título
                ws.merge_cells('A1:H1')
                ws['A1'] = f"Relatório de Produtos - {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
                ws['A1'].font = title_font
                ws['A1'].alignment = center_alignment
                
                # Cabeçalhos
                headers = ['Código', 'Nome', 'Categoria', 'Fornecedor', 'Estoque Atual', 'Estoque Mínimo', 'Preço Custo', 'Preço Venda', 'Status']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                
                # Dados dos produtos
                produtos = Produto.objects.filter(ativo=True).select_related('categoria', 'fornecedor')
                
                for row, produto in enumerate(produtos, 4):
                    status = "Sem Estoque" if produto.estoque_atual == 0 else "Baixo" if produto.estoque_atual <= produto.estoque_minimo else "OK"
                    
                    ws.cell(row=row, column=1, value=produto.codigo)
                    ws.cell(row=row, column=2, value=produto.nome)
                    ws.cell(row=row, column=3, value=produto.categoria.nome)
                    ws.cell(row=row, column=4, value=produto.fornecedor.nome)
                    ws.cell(row=row, column=5, value=produto.estoque_atual)
                    ws.cell(row=row, column=6, value=produto.estoque_minimo)
                    ws.cell(row=row, column=7, value=float(produto.preco_custo))
                    ws.cell(row=row, column=8, value=float(produto.preco_venda))
                    ws.cell(row=row, column=9, value=status)
                    
                    # Aplicar bordas
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).border = border
                
                # Ajustar largura das colunas
                column_widths = [15, 35, 20, 25, 15, 15, 15, 15, 15]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
                
                filename = f"produtos_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
                
            elif tipo_relatorio == 'vendas':
                ws.title = "Vendas"
                
                # Título
                ws.merge_cells('A1:F1')
                ws['A1'] = f"Relatório de Vendas - {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
                ws['A1'].font = title_font
                ws['A1'].alignment = center_alignment
                
                # Cabeçalhos
                headers = ['Data', 'Número Venda', 'Cliente', 'Total', 'Status', 'Vendedor']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                
                # Dados das vendas
                vendas = Venda.objects.filter(
                    data_venda__date__gte=data_inicio,
                    data_venda__date__lte=data_fim
                ).select_related('vendedor').order_by('-data_venda')
                
                for row, venda in enumerate(vendas, 4):
                    ws.cell(row=row, column=1, value=venda.data_venda.strftime('%d/%m/%Y %H:%M'))
                    ws.cell(row=row, column=2, value=venda.numero_venda)
                    ws.cell(row=row, column=3, value=venda.cliente)
                    ws.cell(row=row, column=4, value=float(venda.total))
                    ws.cell(row=row, column=5, value=venda.status)
                    ws.cell(row=row, column=6, value=venda.vendedor.username)
                    
                    # Aplicar bordas
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).border = border
                
                # Ajustar largura das colunas
                column_widths = [20, 20, 30, 15, 15, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
                
                filename = f"vendas_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
                
            elif tipo_relatorio == 'resumo':
                ws.title = "Resumo Geral"
                
                # Título
                ws.merge_cells('A1:D1')
                ws['A1'] = f"Resumo Geral do Estoque - {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
                ws['A1'].font = title_font
                ws['A1'].alignment = center_alignment
                
                # Dados do resumo
                total_produtos = Produto.objects.filter(ativo=True).count()
                produtos_sem_estoque = Produto.objects.filter(ativo=True, estoque_atual=0).count()
                produtos_estoque_baixo = Produto.objects.filter(ativo=True, estoque_atual__lte=F('estoque_minimo')).count()
                
                valor_total_estoque = sum(
                    produto.estoque_atual * produto.preco_custo 
                    for produto in Produto.objects.filter(ativo=True)
                )
                
                movimentacoes_periodo = MovimentacaoEstoque.objects.filter(
                    data_movimentacao__date__gte=data_inicio,
                    data_movimentacao__date__lte=data_fim
                ).count()
                
                # Adicionar dados do resumo
                resumo_data = [
                    ['Métrica', 'Valor'],
                    ['Total de Produtos', total_produtos],
                    ['Produtos Sem Estoque', produtos_sem_estoque],
                    ['Produtos com Estoque Baixo', produtos_estoque_baixo],
                    ['Valor Total do Estoque', f"R$ {valor_total_estoque:.2f}"],
                    ['Movimentações no Período', movimentacoes_periodo],
                ]
                
                for row, (metrica, valor) in enumerate(resumo_data, 3):
                    ws.cell(row=row, column=1, value=metrica).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=valor)
                    
                    # Aplicar bordas
                    for col in range(1, 3):
                        ws.cell(row=row, column=col).border = border
                
                # Ajustar largura das colunas
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                
                filename = f"resumo_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
            
            # Criar resposta HTTP
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Salvar workbook
            wb.save(response)
            
            messages.success(request, f'Relatório {tipo_relatorio} exportado com sucesso!')
            return response
        else:
            messages.error(request, 'Erro no formulário. Verifique os dados e tente novamente.')
    else:
        form = RelatorioDataForm()
    
    return render(request, 'estoque/exportar_relatorio.html', {'form': form})


@login_required
def dashboard_ajax(request):
    """Endpoint AJAX para atualizar dados do dashboard"""
    from django.http import JsonResponse
    
    # Dados atualizados
    produtos_estoque_baixo = Produto.objects.filter(
        ativo=True,
        estoque_atual__lte=F('estoque_minimo')
    ).count()
    
    produtos_sem_estoque = Produto.objects.filter(
        ativo=True,
        estoque_atual=0
    ).count()
    
    total_produtos = Produto.objects.filter(ativo=True).count()
    
    valor_total_estoque = sum(
        produto.estoque_atual * produto.preco_custo 
        for produto in Produto.objects.filter(ativo=True)
    )
    
    # Movimentações dos últimos 7 dias
    movimentacoes_por_dia = []
    for i in range(7):
        data = timezone.now() - timedelta(days=i)
        entrada = MovimentacaoEstoque.objects.filter(
            tipo='ENTRADA',
            data_movimentacao__date=data.date()
        ).count()
        saida = MovimentacaoEstoque.objects.filter(
            tipo='SAIDA',
            data_movimentacao__date=data.date()
        ).count()
        
        movimentacoes_por_dia.append({
            'dia': data.strftime('%d/%m'),
            'entrada': entrada,
            'saida': saida
        })
    
    movimentacoes_por_dia.reverse()
    
    return JsonResponse({
        'produtos_estoque_baixo': produtos_estoque_baixo,
        'produtos_sem_estoque': produtos_sem_estoque,
        'total_produtos': total_produtos,
        'valor_total_estoque': float(valor_total_estoque),
        'movimentacoes_por_dia': movimentacoes_por_dia,
        'timestamp': timezone.now().isoformat()
    })